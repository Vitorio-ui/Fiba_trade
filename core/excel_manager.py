import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from utils.logger import setup_logger

logger = setup_logger()

class ExcelManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self._validate_file_path()
        self.wb = openpyxl.load_workbook(file_path)
        self.sheet = self.wb["Litvinoff"]
        logger.info(f"Excel файл загружен: {file_path}")
    
    def _validate_file_path(self):
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Файл не найден: {self.file_path}")
        if not self.file_path.endswith('.xlsx'):
            raise ValueError("Файл должен быть в формате .xlsx")

    def update_header(self, usdt_balance, deposit):
        """Обновляет шапку с балансами"""
        self.sheet["H1"] = float(usdt_balance)
        self.sheet["H2"] = float(deposit)
        logger.debug(f"Обновлены балансы: USDT={usdt_balance}, Депозит={deposit}")

    def get_active_tickers(self):
        """Получает список активных тикеров"""
        active_tickers = set()
        for row in range(6, self.sheet.max_row + 1):
            status = self.sheet[f"B{row}"].value
            if status == "в работе":
                ticker = self.sheet[f"C{row}"].value
                if ticker:
                    active_tickers.add(ticker)
        return list(active_tickers)

    def update_prices(self, prices):
        """Обновляет текущие цены"""
        updated = 0
        for row in range(6, self.sheet.max_row + 1):
            ticker = self.sheet[f"C{row}"].value
            status = self.sheet[f"B{row}"].value
            if status == "в работе" and ticker in prices:
                self.sheet[f"D{row}"] = float(prices[ticker])
                updated += 1
        logger.debug(f"Обновлено цен: {updated}")

    def simulate_order(self, order_type, row, price, quantity):
        """Симуляция ордера для тестового режима"""
        if order_type == "BUY":
            self.sheet[f"L{row}"] = "Купить"
            self.sheet[f"Q{row}"] = "Открыта"
            self.sheet[f"R{row}"] = "Спот"
        elif order_type == "SELL":
            self.sheet[f"L{row}"] = "Продать"
            self.sheet[f"Q{row}"] = "Закрыта"
        
        self.sheet[f"M{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sheet[f"N{row}"] = float(price)
        self.sheet[f"O{row}"] = float(quantity)
        self.sheet[f"P{row}"] = float(price) * float(quantity)
        self.sheet[f"S{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.sheet[f"T{row}"] = float(price)
        self.sheet[f"U{row}"] = float(quantity)
        self.sheet[f"V{row}"] = float(price) * float(quantity)
        
        logger.info(f"[SIMULATED] Ордер {order_type} в строке {row}: {quantity}@{price}")

    def save(self):
        """Сохраняет изменения в файл"""
        self.wb.save(self.file_path)
        logger.debug("Файл Excel сохранён")